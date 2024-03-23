import tempfile
from datetime import datetime
from io import BytesIO
from typing import List

from fastapi import Depends

from fastapi_pagination import Page, Params
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.workbook import Workbook
from starlette.responses import StreamingResponse

from src.app.worship.daos.songDAO import SongDAO
from src.app.worship.models.database.models import Song
from src.app.worship.factories.songFactory import SongFactory
from src.app.worship.models.songs import ViewSong, CreateAuthor, ViewAuthor, CreateSong


class NoSongException (Exception):
    pass


class SongWithThatCodeExists (Exception):
    pass


class NoAuthorException(Exception):
    pass


class RequiredColumnsNotPresentException(Exception):
    pass


class SongService:
    def __init__(self, song_factory: SongFactory = Depends(SongFactory), song_DAO: SongDAO = Depends(SongDAO)):
        self.song_factory = song_factory
        self.song_DAO = song_DAO

    def get_all_songs(self, params: Params = Params(page=1, size=100)) -> Page[ViewSong]:
        song_entities = []
        song_entities_page = self.song_DAO.get_all_songs(params)
        if song_entities_page:
            for song_entity in song_entities_page.items:
                song_entities.append(self.song_factory.create_song_from_song_entity(song_entity))

            return Page.create(items=song_entities, params=params, total=song_entities_page.total)
        else:
            return []

    def get_song_by_id(self, id) -> ViewSong:
        song_entity = self.song_DAO.get_song_by_id(id)
        if song_entity is None:
            raise NoSongException("Song with that id does not exist")
        return self.song_factory.create_song_from_song_entity(song_entity)

    def create_song(self, song: CreateSong) -> ViewSong:
        #check if song with that code exists and if so raise error
        if self.song_DAO.get_song_by_code(song.code):
            raise SongWithThatCodeExists("Song with that code already exists")
        valid_authors = []
        if song.author_ids:
            for author_id in song.author_ids:
                author = self.song_DAO.get_author_by_id(author_id)
                if author:
                    valid_authors.append(author)
                else:
                    raise NoAuthorException("Author with that id does not exist")
        song_entity = self.song_factory.create_song_entity_from_song(song, valid_authors)
        song_entity = self.song_DAO.create_song(song_entity)
        return self.song_factory.create_song_from_song_entity(song_entity)

    def create_author(self, author: CreateAuthor) -> ViewAuthor:
        author_entity = self.song_factory.create_author_entity_from_author(author)
        author_entity = self.song_DAO.create_author(author_entity)
        return self.song_factory.create_author_from_author_entity(author_entity)

    def update_author(self, id, author):
        author_entity = self.song_DAO.get_author_by_id(id)
        if author_entity is None:
            raise NoAuthorException("Author with that id does not exist")
        update_values = author.dict(exclude_unset=True)
        author_entity = self.song_DAO.update_author(author_entity, update_values)
        return self.song_factory.create_author_from_author_entity(author_entity)

    def get_all_authors(self):
        author_entities = self.song_DAO.get_all_authors()
        authors = []
        for author_entity in author_entities:
            authors.append(self.song_factory.create_author_from_author_entity(author_entity))
        return authors

    def get_author_by_id(self, id):
        author_entity = self.song_DAO.get_author_by_id(id)
        if author_entity is None:
            raise NoAuthorException("Author with that id does not exist")
        return self.song_factory.create_author_from_author_entity(author_entity)

    def update_song(self, id, song):
        song_entity = self.song_DAO.get_song_by_id(id)
        if song_entity is None:
            raise NoSongException("Song with that id does not exist")
        update_values = song.dict(exclude_unset=True)

        authors_ids = update_values.pop('author_ids', song.dict())
        #validate author_ids
        valid_authors = []
        if authors_ids:
            for author_id in authors_ids:
                author = self.song_DAO.get_author_by_id(author_id)
                if author:
                    valid_authors.append(author)
                else:
                    raise NoAuthorException("Author with that id does not exist")

        if authors_ids:
            existing_author_songs = self.song_DAO.get_existing_authors_for_song(song_entity.id)
            for existing_author_song in existing_author_songs:
                self.song_DAO.delete_author_song(existing_author_song.id)

            for author in valid_authors:
                self.song_DAO.create_author_song(song_entity, author)

        song_entity = self.song_DAO.update_song(song_entity, update_values)
        return self.song_factory.create_song_from_song_entity(song_entity)

    def upload_songs(self, file, force_song_creation_with_same_name):
        # parse excel sheet. It will have a row with title of the columns. Some columns are optional, and some
        # are required. Optional are code, song_key, style, tempo, ccli_number, video_link, secondary_name.
        # Required are name and author. Author is a comma separated list of author names.

        # The logic should be to iterate through the authors and create them if they don't exist. Then create the song
        # and associate the authors with the song. If the song already exists, update it. If the author already exists,
        # The only way to be certain that a song already exists is if the code is provided. If the code is not provided
        # then we will assume that the song does not exist and create it. If the code is provided, then we will check
        # if the song exists and if it does, we will update it.

        # start processing the xlsx file, please stop using comments and start implementing the method
        songs_created = 0
        songs_updated = 0
        authors_created = 0
        workbook = load_workbook(file.file)
        sheet = workbook.active
        # get column indexes from first row
        column_indexes = {}
        for i, cell in enumerate(sheet[1]):
            column_indexes[str(cell.value).upper()] = i

        print(column_indexes);

        # validate that required columns are present
        required_columns = ['NAME', 'AUTHOR']
        for required_column in required_columns:
            if required_column not in column_indexes:
                raise RequiredColumnsNotPresentException(f"Column {required_column} is required")

        # iterate through author column rows and create authors if they don't exist
        author_column_index = column_indexes['AUTHOR']
        author_column = sheet.iter_cols(min_col=author_column_index+1, max_col=author_column_index+1, values_only=True)
        #skip the first row with column names

        author_names = [cell for tuple in author_column for cell in tuple]
        #skip the first row with column names
        author_names = author_names[1:]
        for author_name_list in author_names:
            #split the author names by comma
            if author_name_list is not None and author_name_list != "":
                author_names = author_name_list.split(',')
                for author_name in author_names:
                    if author_name:
                        author_name = author_name.strip()
                        #process if author name is not a blank or empty string
                        if author_name != "":
                            # note this may create duplicate authors if existing names are slightly different.
                            author = self.song_DAO.get_author_by_name(author_name)
                            if author is None:
                                author = self.song_DAO.create_author(self.song_factory.create_author_entity_from_author(CreateAuthor(name=author_name)))
                                authors_created += 1
                                print (f"Created author {author.name}")

        # iterate through the rows and create or update songs
        for row in sheet.iter_rows(min_row=2, values_only=True):
            # get the song code if it exists
            code = None
            if 'CODE' in column_indexes:
                code = row[column_indexes['CODE']]

            if code:
                song = self.song_DAO.get_song_by_code(code)
                if song:
                    # update the song
                    song = self.song_factory.create_song_from_song_entity(song)
                    print(f"Updating song {song.name}")
                    # create the CreateSong object but only specify columns if they exist in the spreadsheet
                    create_song = self.create_song_from_row(column_indexes, row)
                    print(create_song.dict())
                    self.update_song(song.id, create_song)
                    songs_updated += 1
                else:
                    # create the song
                    print(f"Creating song {row[column_indexes['NAME']]}")
                    create_song = self.create_song_from_row(column_indexes, row)

                    print(create_song.dict())
                    self.create_song(create_song)
                    songs_created += 1
            else:
                # create the song
                create_new_song = True
                if not force_song_creation_with_same_name:
                    # try to find the song by name
                    song = self.song_DAO.get_song_by_name(row[column_indexes['NAME']])
                    if song:
                        create_new_song = False

                create_song = self.create_song_from_row(column_indexes, row)
                if create_new_song:
                    print(f"Creating new song {row[column_indexes['NAME']]}")

                    print(create_song.dict())
                    self.create_song(create_song)
                    songs_created += 1
                else:
                    # update the song
                    print(f"Updating song {song.name}")
                    # create the CreateSong object but only specify columns if they exist in the spreadsheet
                    create_song = self.create_song_from_row(column_indexes, row)
                    print(create_song.dict())
                    song = self.song_DAO.get_song_by_name(row[column_indexes['NAME']])
                    self.update_song(song.id, create_song)
                    songs_updated += 1

        return "Songs created: " + str(songs_created) + " Songs updated: " + str(songs_updated) + " Authors created: " + str(authors_created)

    def export_songs_as_xlsx(self):
        # get all songs from the database
        songs = self.song_DAO.get_all_songs_without_pagination()
        # crete new workbook
        workbook = Workbook()
        sheet = workbook.active
        # write the column names in the first row
        bold_font = Font(bold=True)
        fill = PatternFill(start_color="C0C0FF", end_color="C0C0FF", fill_type="solid")

        #sheet.append(['CODE', 'NAME', 'SECONDARY_NAME', 'SONG_KEY', 'STYLE', 'TEMPO', 'CCLI_NUMBER', 'VIDEO_LINK', 'AUTHOR'])
        headers = ['CODE', 'NAME', 'SECONDARY_NAME', 'SONG_KEY', 'STYLE', 'TEMPO', 'CCLI_NUMBER', 'VIDEO_LINK',
                   'AUTHOR']
        for i, header in enumerate(headers, start=1):
            cell = sheet.cell(row=1, column=i, value=header)
            cell.font = bold_font
            cell.fill = fill

        for song in songs:
            author_names = []
            for author_song in song.authors:
                author_names.append(author_song.author.name)
            sheet.append([song.code, song.name, song.secondary_name, song.song_key, song.style, song.tempo, song.ccli_number, song.video_link, ', '.join(author_names)])

        for column in sheet.columns:
            max_length = 0
            column = [cell for cell in column]
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2)
            sheet.column_dimensions[column[0].column_letter].width = adjusted_width

        stream = BytesIO()
        workbook.save(stream)
        stream.seek(0)

        # Return a StreamingResponse to stream the file back to the client
        return StreamingResponse(stream, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                 headers={'Content-Disposition': f'attachment; filename=songs-{datetime.now().strftime("%Y%m%d-%H%M%S")}.xlsx'})

    def create_song_from_row(self, column_indexes, row):
        create_song = CreateSong()
        for key in column_indexes.keys():
            if str(key).lower() in CreateSong.__annotations__:
                create_song.__setattr__(str(key).lower(), row[column_indexes.get(key)])
        # get list of author id's from row
        author_ids: List[int] = self.get_author_id_list(column_indexes, row)
        create_song.author_ids = author_ids
        return create_song

    def get_author_id_list(self, column_indexes, row):
        author_ids = []
        if 'AUTHOR' in column_indexes:
            if row[column_indexes['AUTHOR']] is None or row[column_indexes['AUTHOR']] == "":
                return author_ids
            author_names = row[column_indexes['AUTHOR']].split(',')
            for author_name in author_names:
                author = self.song_DAO.get_author_by_name(author_name.strip())
                if author:
                    author_ids.append(author.id)
        return author_ids
